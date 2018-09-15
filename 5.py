# -*- coding: utf-8 -*-
import time
import os 
import openpyxl
from collections import Counter
from openpyxl.styles import Font,Alignment
from functools import reduce

#需要整理的excel路径写在这里
abspath='C:\\Users\\admin\\合肥运维账号权限清单_20180113.xlsx'
#abspath='C:\\Users\\admin\\1.xlsx'

#各个系统的表头生成函数
def excel(excel_name):
	wb1=openpyxl.Workbook()
	#设置字体类型
	font_head=Font(u'仿宋_GB2312',size=20,bold=True)
	#打开工作表的所有sheet
	all_sheet=wb.sheetnames
	#设置位置
	align = Alignment(horizontal='center', vertical='center')
	#生成表头文件
	sheet=wb1['Sheet']
	sheet['A1']=excel_name+'运维账号明细表'
	sheet['A1'].font=font_head
	sheet['A1'].alignment=align
	sheet['A2']='编号:'
	sheet['A3']='最后更新日期: 脚本运行时间'
	sheet['E3']='最后更新人:'
	sheet['A4'],sheet['B4'],sheet['C4'],sheet['D4']='系统名称','主机名','是否重要系统','账号名称'
	sheet['E4'],sheet['F4'],sheet['G4'],sheet['H4']='密码掌管人','使用人','账号类型','用途说明'
	sheet.merge_cells('A2:H2')
	sheet.merge_cells('A1:H1')
	sheet.merge_cells('A3:D3')
	sheet.merge_cells('E3:H3')
	sheet.print_options.horizontalCentered = True      #水平居中 
	print(sheet['A1'].value)
	#根据要求设置行高
	sheet.row_dimensions[1].height=25
	sheet.row_dimensions[2].height=18
	sheet.row_dimensions[3].height=18
	sheet.row_dimensions[4].height=25
	#根据要求设置列高
	sheet.column_dimensions['A'].width=20
	sheet.column_dimensions['B'].width=12
	sheet.column_dimensions['C'].width=12
	sheet.column_dimensions['D'].width=12
	sheet.column_dimensions['E'].width=12
	sheet.column_dimensions['F'].width=12
	sheet.column_dimensions['G'].width=12
	sheet.column_dimensions['H'].width=20
	#将A1输入数值并且保存这个工作表为C:\\Users\\admin\\test.xlsx
	i='D:\\tmp\\'+excel_name+'运维账号明细表.xlsx'
	wb1.save(i)


#打开这个工作表
wb=openpyxl.load_workbook(abspath)#打开工作表
#读第一个表的相关数据
work_sheet=wb['Sheet3']
max_column=work_sheet.max_column#这个sheet的最大列
max_row=work_sheet.max_row#这个sheet的最大行
#打印这个工作表的最大行和最大列
#print(max_row,max_column)
#整体函数部分
#外部样式文件
align = Alignment(horizontal='center', vertical='center')

d={}
l1=[]
for i in range(1,max_row+1):
	l=[]
	l.append(work_sheet.cell(row=i,column=1).value,)
	l.append(work_sheet.cell(row=i,column=2).value,)
	l.append(work_sheet.cell(row=i,column=3).value,)
	l.append(work_sheet.cell(row=i,column=4).value,)
	l.append(work_sheet.cell(row=i,column=5).value,)
	l.append(work_sheet.cell(row=i,column=6).value,)
	d[i]=l
for v in d.values():
	l1.append(v)
l1.pop(0)
#print(l1)
#获取系统个数以及每个系统的主机数
l2,l3,l4=[],[],[]
for i in l1:
	l4.append(i[0])
	if i[0] not in l2:
		l2.append(i[0])
#生成所有系统与主机的键值对
d1=Counter(l4).most_common(len(l4))
d1=dict(d1)
print('这是测试每个系统有多少个主机的 %s' %d1)
#l1=[ 
#    ['CAAS管理平台', 'caas-hf-biz-nc19', '否', 'root', '特权', '张海\\侯登永'],
#    ['CAAS管理平台', 'caas-hf-ebk-nc41', '否', 'root', '特权', '张海\\侯登永'],
#   ]

#l2是所有系统的个数
print('这是L2 %s' %(l2))
#定义第二列的控制函数
list_func1={'0':0}
def function2(x):
	global list_func1
	list1=[j[1] for i in l2 for j in l1 if i==j[0]]
	#print(list_func4)
	if x not in list_func1.keys():
		list_func1.update({x:d1.get(x)})
	list_control=[]
	for i in list_func1.values():
		list_control.append(i)
	a,b=reduce(lambda x,y:x+y,list_control[0:len(list_control)-1]),list_control[-1]
	list=list1[a:a+b]
	return list
#定义第四列的控制函数
list_func2={'0':0}
def function3(x):
	global list_func2
	list1=[j[3] for i in l2 for j in l1 if i==j[0]]
	#print(list_func4)
	if x not in list_func2.keys():
		list_func2.update({x:d1.get(x)})
	list_control=[]
	for i in list_func2.values():
		list_control.append(i)
	a,b=reduce(lambda x,y:x+y,list_control[0:len(list_control)-1]),list_control[-1]
	list=list1[a:a+b]
	return list

#定义第五六列的控制函数
list_func3={'0':0}
def function5_6(x):
	global list_func3
	list1=[j[5] for i in l2 for j in l1 if i==j[0]]
	#print(list_func4)
	if x not in list_func3.keys():
		list_func3.update({x:d1.get(x)})
	list_control=[]
	for i in list_func3.values():
		list_control.append(i)
	a,b=reduce(lambda x,y:x+y,list_control[0:len(list_control)-1]),list_control[-1]
	list=list1[a:a+b]
	return list

#定义第七列的控制函数,循环生成每个系统里面的主机名的列表
list_func4={'0':0}
def function7(x):
	global list_func4
	list1=[j[4] for i in l2 for j in l1 if i==j[0]]
	#print(list_func4)
	if x not in list_func4.keys():
		list_func4.update({x:d1.get(x)})
	list_control=[]
	for i in list_func4.values():
		list_control.append(i)
	a,b=reduce(lambda x,y:x+y,list_control[0:len(list_control)-1]),list_control[-1]
	list=list1[a:a+b]
	return list

#第八列匹配字典
d_8={
 r'特权':'系统维护',
 r'操作':'应用维护',
 r'维护':'应用维护',
 r'只读':'系统查询',
}

#列表的第一项内容生成以及格式规范
#print([j[1] for i in l2 for j in l1 if i==j[0]])
#print(len([j[1] for i in l2 for j in l1 if i==j[0]]))

for i in l2:
	excel(i)#生成各个表头文件以及表的名字
	abspath='D:\\tmp\\'+i+'运维账号明细表.xlsx'
	#print('这是测试abspath %s'%abspath)
	wb2=openpyxl.load_workbook(abspath)
	work_sheet=wb2['Sheet']
	#print('这里是测试 %s'% j[1])
	for k in range(0,d1.get(i)):
		work_sheet.cell(row=k+5,column=1).value=i
		work_sheet.cell(row=k+5,column=2).value=function2(i)[k]
		work_sheet.cell(row=k+5,column=3).value='否'
		work_sheet.cell(row=k+5,column=4).value=function3(i)[k]
		work_sheet.cell(row=k+5,column=5).value=function5_6(i)[k]
		work_sheet.cell(row=k+5,column=6).value=function5_6(i)[k]
		work_sheet.cell(row=k+5,column=7).value=function7(i)[k]
		work_sheet.cell(row=k+5,column=8).value=d_8.get(work_sheet.cell(row=k+5,column=7).value)
	work_sheet.merge_cells('A5:A'+str(d1.get(i)+4))
	#abspath是所有的excel生成路径
	abspath1='D:\\tmp\\'+i+'运维账号明细表.xlsx'
	work_sheet['A5'].alignment=align
	wb2.save(abspath1)



			





















		


